#do a progress bar on the display screen

from createobjects import continuingStudentObjects
from createobjects import freshmenObjects
import pandas as pd
import openpyxl
import os

listofcontinuingstudents = continuingStudentObjects()
listoffreshmen = freshmenObjects()

print(len(listofcontinuingstudents), len(listoffreshmen))

'''
    @brief: pairing constant determines if the number of continuing student present and their 
    cardinality guarantees the pairing of all male and female freshmen. 

    @return: true if value is greater than 1 else false. If it is 'True' then they are as many or more 
    continuing females/males for the fresher females/males and vice versa
'''

def pairingConstant(gender):
    
    if gender=="Male" or gender=="Female":
        sumcon =0

        for c in listofcontinuingstudents:
            if(c.getGender()==gender):
                sumcon+=c.getCardinality()

        sumfres =0

        # count male/females in incoming class.     
        for f in listoffreshmen:
            if(f.getGender()==gender[0]):
                sumfres+=1

        try:
            result = sumcon/sumfres
            if(result > 1):
                return True
            else:
                return False 
        except ZeroDivisionError as e:
            print("Error: Cannot divide by zero")
    else:
        return "wrong parameter passed [in pairingConstant module]" 
        

    
# Ensure that Ghana is the last option in the list of preferred nationalities 
def moveGhana(continuingstudent):
    list = continuingstudent.getPreferredNationality()
    if len(list)>1 and "Ghana" in list:
        list.remove("Ghana")
        list.insert(len(list), "Ghana") #insert at the end of the list
    continuingstudent.setPreferredNationality(list)
    return continuingstudent

# Function to return the pairing description when pairing has been done.
def pairDescription(number, continuingstudent, pairsummary, pair):
    pairsummary["continuingstudent"] = continuingstudent
    pairsummary["freshers"] = pair

    if continuingstudent.getCardinality()==len(pair): # complete pairing
        listofcontinuingstudents.remove(continuingstudent)
        pairsummary["status"] = "complete"
    else: #incomplete pairing
        pairsummary["status"] = "incomplete"
    return pairsummary
    

def pairing(continuingstudent, freshmen, exceptcountry):
    pair = []
    pairsummary ={}

    # Base pairing: continuing students wants x number of freshers from the same nationality, without exceptions ("None") AND
    # secondary pairing: continuing student wants x number of freshers from N different nationalities without exceptions

    if exceptcountry[0]=="None":
        # Ensure that Ghana is the last option in the list of preferrednationalities if it was specified.
        continuingstudent= moveGhana(continuingstudent)

        for i in range(continuingstudent.getCardinality()):
        
            if len(continuingstudent.getPreferredNationality())==1:
                for f in freshmen:
                    # pairing when "Any" is in nationality
                    if "Any" in continuingstudent.getPreferredNationality()[0]:
                        if continuingstudent.getGender()[0]== f.getGender()[0]:
                            pair.append(f)
                            listoffreshmen.remove(f) # remove paired fresher 
                            break 

                    # pairing when countries specified                    
                    elif f.getNationality() in continuingstudent.getPreferredNationality() and continuingstudent.getGender()[0]== f.getGender():
                        pair.append(f)
                        listoffreshmen.remove(f) # remove paired fresher 
                        break
                    else:
                        continue
            elif len(continuingstudent.getPreferredNationality())>1:
                for f in freshmen: 
                    if f.getNationality() in continuingstudent.getPreferredNationality() and continuingstudent.getGender()[0]== f.getGender():
                        pair.append(f)
                        listoffreshmen.remove(f)
                        break
                    else:
                        continue
            
            if i==continuingstudent.getCardinality()-1: # condition checks if pairing is complete
                pairsummary = pairDescription(i, continuingstudent, pairsummary, pair)
                return pairsummary
            else: 
                continue           

    #  it has country(ies) not 'None'
    elif exceptcountry[0]!="None": 
        index = 0
        for i in range(continuingstudent.getCardinality()):
            country = exceptcountry[index]
            for f in freshmen:
                # don't pair with a fresher of the 'except nationality' 
                if country!=f.getNationality() and continuingstudent.getGender()[0]== f.getGender():
                    pair.append(f) 
                    listoffreshmen.remove(f) # remove paired fresher
                    break
                else:
                    continue

            if i==continuingstudent.getCardinality()-1: # condition checks if pairing is complete
                pairsummary = pairDescription(i, continuingstudent, pairsummary, pair)
                return pairsummary
            else: 
                # if continuing student specified more than one 'except country' so we can increase the index.
                if len(exceptcountry)-index>1: 
                    index+=1
                    continue
                else:
                    #keep the index unchanged
                    continue              
    else:  
        return 404 #error message


class Edges:    

    def __init__(self, listofcontinuingstudents, listoffreshmen, matched, reserved):
        self.listofcontinuingstudents = listofcontinuingstudents
        self.listoffreshmen = listoffreshmen
        self.matched = matched
        self.reservedpairs =reserved #continuing students who are not paired at the end of the matching.

    def getMatched(self):
        return self.matched
    
    def getReserved(self):
        return self.reservedpairs
    
    def setMatched(self, matched):
        self.matched = matched 

    def setReserved(self, reserved):
        self.reservedpairs=reserved
    
    def matching(self):

        # first check if they're enough male and female continuing students to be paired with male freshers
        if pairingConstant("Male")==True:
            if pairingConstant("Female")==True:
                countpairs=0
                countcardinality =0
                countcountinuingstudents = 0
                notpaired =0
                for c in listofcontinuingstudents:
                    paired=dict()
                    exceptcountry = []            
                    pnationality = c.getPreferredNationality()
                    for country in pnationality:
                        # first check if 'not country' exist
                        if "Not" in country:
                            exceptcountry.append(country.split("Not")[1])
                            break
                        elif "Any" in country: 
                            exceptcountry.append("None")
                            break
                        else: 
                            exceptcountry.append("None") # a case where we have actual countries.

                    # Commence pairing
                    if(len(listoffreshmen)==0):
                        print("Hurray, match is complete")    
                        print(len(listofcontinuingstudents), len(listoffreshmen))
                        print("\n\n","number of matched pairs: ",countpairs,"\n total cardinalities used", countcardinality, "\n paired countinuing students",countcountinuingstudents, "\n not paired countinuing students", notpaired,"\n\n")                 
                        return self.matched,self.reservedpairs
                    else:
                        summary = pairing(c, listoffreshmen, exceptcountry)  
                    
                    print('\n',summary['continuingstudent'].toString())
                    for i in range(len(summary['freshers'])):
                        print(summary['freshers'][i].toString())

                    #**
                    # this part of the code is getting too big, create another function for this
                    # next this is the part that would be useful for writing paired information to dataframe.
                    #**

                    # a pairing summary status may or may not be complete, 
                    # if status is complete, append the paired item into the list.
                    if summary["status"]=="complete":
                        paired["continuingstudent"] = summary["continuingstudent"]
                        paired["freshers"] = summary["freshers"]
                        self.matched.append(paired)
                        countcardinality+=c.getCardinality()
                        countcountinuingstudents+=1
                        countpairs+=1
                        print('I am complete')
                    elif summary["status"] == "incomplete":
                        paired["continuingstudent"] = summary["continuingstudent"]
                        paired["freshers"] = summary["freshers"]
                        if(len(summary["freshers"])==0):
                            print('I am incomplete: I don\'t have any of my freshers')
                            self.reservedpairs.append(paired['continuingstudent'])
                            notpaired+=1
                        else:
                            print('I am incomplete: I don\'t have all of my freshers')
                            listofcontinuingstudents.remove(c)
                            countcardinality+=c.getCardinality()
                            countpairs+=1
                            countcountinuingstudents+=1
                            self.matched.append(paired)
                    else: 
                        continue
            else:
                print("Error: Not enough females for pairing, repopulate female cardinality and then do matching again.") 
                return self.matched,self.reservedpairs 
        else: 
            print("Error: Not enough males for pairing, repopulate male cardinality and then do matching again.")   
            return self.matched,self.reservedpairs   


matched = []
reserved = []
E = Edges(listofcontinuingstudents, listoffreshmen, matched, reserved)

def generateExcelFiles(status="paired-list", edges=E):
    matched, reservedpairs = edges.matching()

    print(len(matched), len(reservedpairs))

    if not os.path.exists('../controller/downloads'):
        os.makedirs('../controller/downloads')

    # Initialize an empty list to store data
    data = []

    if(status=="paired-list"):
        # Loop through each pair of matched students
        for match in matched:
            cs = match['continuingstudent']
            for f in match['freshers']:
                # Create a dictionary with the data for this pair
                record = {
                    'freshername': f.getName(),
                    'buddyname': cs.getName(),
                    'buddynationality': 'empty',
                    'buddygender': cs.getGender(),
                    'buddyyeargroup': 'empty',
                    'buddyemail': 'empty',
                    'buddyphonenumber': 'empty',
                    'funfact': 'empty'

                }
                # Add this dictionary to the list of data
                data.append(record)

        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Write the header row to the sheet
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header

        # Write the data to the sheet
        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = record[header]

        # Save the workbook to a file
        wb.save('../controller/downloads/matched.xlsx')

        return "Successfully generated paired list"
    
    elif(status=="unpaired-list"):
        delimiter = ','

        # Loop through each pair of matched students
        for cs in reservedpairs:
            # Create a dictionary with the data for this pair
            record = {
                'Fullname': cs.getName(),
                'Gender': cs.getGender(),
                'Nationality': 'empty',            
                'PreferredNationality': delimiter.join(cs.getPreferredNationality()),
                'NumberofFreshers':cs.getCardinality(),
                'Email': 'empty',
                'Phonenumber': 'empty',
                'Funfact': 'empty'
            }
            # Add this dictionary to the list of data
            data.append(record)
        
        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Write the header row to the sheet
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header

        # Write the data to the sheet
        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = record[header]

        # Save the workbook to a file
        wb.save('../controller/downloads/unmatched.xlsx')
        return "Successfully generated list of unpaired continuing students"
    else:
        return "Internal error"


generateExcelFiles(status="paired-list", edges=E)
generateExcelFiles(status="unpaired-list",edges=E)